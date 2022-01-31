import warnings
import os, sys, glob
import numpy as np
import pandas as pd
import multiprocessing
import xlrd
from openpyxl import load_workbook
import seaborn as sns
from matplotlib import pyplot as plt
from PyQt5 import QtGui, QtWidgets


warnings.simplefilter(action='ignore')
FOLDER = '/home/denis/Documents/Covid/Rosstat/'
COLUMN_REGION = 0
COLUMN_MORTALITY = 5
QUANTITY_REGIONS = 82

    
class Analyzator():
    def __init__(self):
        sns.set_style('darkgrid')
        
    def analyze(self):
        all_files = self._get_excel_files()
        with multiprocessing.Pool() as pool:
            all_data = pool.map(self._read_excel_file, all_files)
        
        # all_data = []
        # for file_name in all_files:
        #     all_data.append(self._read_excel_file(file_name))
            
        data = pd.concat(all_data)
        if not self._data_is_correct(data):
            return 0
        
        popular_regions = data[data.Year<=2019]
        popular_regions = popular_regions.drop(['Year', 'Month'], axis=1)
        grouped = popular_regions.groupby('Region')
        popular_regions = grouped.mean()
        popular_regions = popular_regions.sort_values(by='Mortality', ascending=False)
        popular_regions = set(popular_regions.index[:10])
            
        for region in popular_regions:
            mean_region = data[(data.Region==region) & (data.Year<=2019)]
            mean_region = mean_region.drop(['Region', 'Year'], axis=1)
            grouped = mean_region.groupby('Month')
            mean_region = grouped.mean()
            mean_region = mean_region.rename({'Mortality': 'Mean_Mortality'}, axis=1)
            mean_region['Mean_Mortality'] = np.round(mean_region['Mean_Mortality'])
            
            all_data = None
            for year in (2020, 2021):
                year_data = data[(data.Region==region) & (data.Year==year)]
                year_data = year_data.drop(['Region', 'Year'], axis=1)
                year_data = pd.merge(year_data, mean_region, left_on=['Month'], right_index=True)
                year_data[str(year)] = year_data.Mortality / year_data.Mean_Mortality
                year_data = year_data[['Month', str(year)]]
                
                if all_data is None:
                    all_data = year_data
                else:
                    all_data = pd.merge(all_data, year_data, how='left', on='Month')
            
            plotter = Plotter(region)    
            plotter.plot_bar(all_data)
            plt.show()
        
    def _data_is_correct(self, data):
        grouped = data.groupby(['Year', 'Month'])
        regions = grouped.count()['Region']
        if len(set(regions)) == 1:
            return 1
        
        normal = regions[regions == QUANTITY_REGIONS]
        year_month = normal.index[0]
        normal_regions = set(data.Region[(data.Year==year_month[0]) & (data.Month==year_month[1])])
        
        unnormal = regions[regions != QUANTITY_REGIONS]
        for year, month in unnormal.index:
            extended_regions = set(data.Region[(data.Year==year) & (data.Month==month)])
            starnge_region = extended_regions - normal_regions
            print('starnge_region = ' + starnge_region)
        return 0
        
    def _get_excel_files(self):
        all_files = []    
        pattern_folder = 'edn' + '[0-9]' * 4
        folders = glob.glob(FOLDER + pattern_folder)
        folders = sorted(folders)
        for folder in folders:
            all_files.extend(glob.glob(folder + '/*.xlsx'))
            all_files.extend(glob.glob(folder + '/*.xls'))
        return all_files
    
    @staticmethod
    def _read_excel_file(file_name):
        excel = Excel(file_name)
        return excel.read()


class Plotter:
    def __init__(self, region):
        self.fig = plt.figure()
        self.fig.suptitle(region, fontsize=14)
        
        self.ax = self.fig.add_subplot(1, 1, 1)
        self.ax.set_ylabel('Excess mortality', fontsize=12)
        
        line = plt.axhline(y=1, linewidth=3, color='k')
        self.ax.add_patch(line)
        
    def plot_bar(self, data):
        data = data.sort_values(by='Month')
        months = {1:"Январь", 2:"Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь", \
                  7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"}
        data = data.replace(months)
        data = data.set_index('Month')
        data.plot.bar(ax=self.ax, rot=30)
                    
                    
class Excel:
    def __init__(self, file_name):
        self.file_name = file_name
        
    def read(self):
        if self.file_name.endswith('xls'):
            return self._read_xls()
        elif self.file_name.endswith('xlsx'):
            return self._read_xlsx()
        else:
            raise RuntimeError("Your file- extenstion doesn't support!")
    
    def _read_xls(self):
        data = []
        year = self._get_year()
        month = self._get_month()
        
        work_book = xlrd.open_workbook(self.file_name)
        sheet = work_book.sheet_by_name('t1_1')
        
        row_count = sheet.nrows
        for row in range(6, row_count):
            region = sheet.cell(row, COLUMN_REGION).value
            region = self._correct_region(region)
            if not self._valid_region(region):
                continue
            
            mortality = sheet.cell(row, COLUMN_MORTALITY).value
            if not mortality:
                continue
            
            data.append((year, month, region, int(mortality)))
        return pd.DataFrame(data, columns=['Year', 'Month', 'Region', 'Mortality'])
        
    def _read_xlsx(self):
        data = []
        month = self._get_month()
        year = self._get_year()
            
        work_book = load_workbook(filename=self.file_name, read_only=True)
        sheet = work_book['t1_1']
        for index, row in enumerate(sheet.rows):
            if index <= 5:
                continue
            
            region = row[COLUMN_REGION].value
            region = self._correct_region(region)
            if not self._valid_region(region):
                continue
            
            mortality = row[COLUMN_MORTALITY].value
            if not mortality:
                continue
            
            data.append((year, month, region, int(mortality)))
        return pd.DataFrame(data, columns=['Year', 'Month', 'Region', 'Mortality'])
    
    def _get_year(self):
        base_name = os.path.basename(self.file_name)
        pos1 = base_name.find('201')
        pos1 = pos1 if pos1>=0 else 1000
        pos2 = base_name.find('202')
        pos2 = pos2 if pos2>=0 else 1000
        pos = min(pos1, pos2)
        return int(base_name[pos: pos+4])
    
    def _get_month(self):
        base_name = os.path.basename(self.file_name)
        pos1 = base_name.find('0')
        pos1 = pos1 if pos1>=0 else 1000
        pos2 = base_name.find('1')
        pos2 = pos2 if pos2>=0 else 1000
        pos = min(pos1, pos2)
        return int(base_name[pos: pos+2])
    
    def _correct_region(self, region):
        if region is None:
            return region
        
        region = region.strip()
        
        pos = region.find('(')
        if pos >= 0:
            region = region[:pos]
        
        if region == 'Республика Северная Осетия-Алания':
            region = 'Республика Северная Осетия- Алания'
             
        symbol_table = str.maketrans('', '', '0123456789(){}[]')
        region = region.translate(symbol_table)
        return region
        
    def _valid_region(self, region):
        if region is None:
            return False
        
        if region == 'А' or \
            region.endswith('авт.округ') or \
            region.endswith('без автономии') or \
            region.endswith(' округ') or \
            region == 'Российская Федерация':
            return False
        return True

def analyze():
    counter = Analyzator()
    counter.analyze()     
              
def build_window():       
    window = QtWidgets.QWidget()
    window.resize(600, 100)
    
    label = QtWidgets.QLabel()
    label.setText('Анализ избыточной смертности по месяцам для 10 самых крупных регионов')

    myFont=QtGui.QFont()
    myFont.setPointSize(14)
    myFont.setBold(True)
    label.setFont(myFont)

    button = QtWidgets.QCommandLinkButton("Analyze")
    button.clicked.connect(analyze)
    
    vbox = QtWidgets.QVBoxLayout()
    vbox.addWidget(label)
    vbox.addWidget(button)
    window.setLayout(vbox)
    return window
    
if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = build_window()    
    window.show()
    sys.exit(app.exec_())
            
    
